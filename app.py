import os
import csv
import json
import math
import re
import unicodedata
import asyncio
from dataclasses import dataclass, asdict
from datetime import datetime, timezone, date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin
from io import BytesIO
import hashlib
import streamlit as st
import streamlit.components.v1 as components
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from fastembed import TextEmbedding
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ✅ Fix Playwright subprocess on Windows under Streamlit
if os.name == "nt":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

# -----------------------------
# Constants
# -----------------------------
LANDING_URL_TEMPLATE = "https://monaqasat.mof.gov.qa/TendersOnlineServices/AvailableMinistriesTenders/{page}"
BASE_URL = "https://monaqasat.mof.gov.qa"

APP_ROOT = Path(__file__).resolve().parent
DEFAULT_INTERESTS_PATH = APP_ROOT / "interests.json"
LAST_SCRAPE_FILE = APP_ROOT / ".last_successful_scrape_date.txt"  # stores YYYY-MM-DD
ISIC_DEFAULT_PATH = APP_ROOT / "isic-of-interests.json"

DEFAULT_INTERESTS_FALLBACK = {
  "Digital Transformation": [
    "End-to-end digital transformation programs: portals, e-services, workflow/case management, integration and automation.",
    "التحول الرقمي: بوابات وخدمات إلكترونية وأتمتة إجراءات وإدارة حالات وتكامل بين الأنظمة."
  ],
  "Mobile App Development": [
    "Mobile app development for iOS/Android, Flutter/React Native, API integration, security, support & maintenance.",
    "تطوير تطبيقات الجوال iOS/Android (Flutter/React Native) وتكامل API ودعم وصيانة."
  ],
  "Enterprise Architecture": [
    "Enterprise Architecture consulting, capability maps, target architecture, governance, EA repositories (TOGAF/ArchiMate).",
    "معمارية المؤسسة: خرائط قدرات، معمارية مستهدفة، حوكمة، مستودعات EA (TOGAF/ArchiMate)."
  ],
  "AI Implementations": [
    "Applied AI: chatbots, document AI/OCR, NLP extraction/classification, RAG assistants, analytics & automation.",
    "حلول الذكاء الاصطناعي: شات بوت، فهم مستندات/OCR، NLP، مساعد معرفي، تحليلات وأتمتة."
  ],
  "Call Center Operations": [
    "Contact center operations/outsourcing: omnichannel support, IVR, CRM/ticketing, WFM, QA, SLAs.",
    "تشغيل مراكز الاتصال: دعم متعدد القنوات، IVR، تكامل CRM/تذاكر، إدارة ورديات، جودة، SLAs."
  ]
}

# -----------------------------
# Data models
# -----------------------------
@dataclass
class TenderCard:
    tender_no: str
    title: str
    details_url: str
    tender_id: Optional[str] = None
    publish_date: Optional[str] = None
    requested_sector_type: Optional[str] = None
    tender_bond_qar: Optional[str] = None
    documents_value_qar: Optional[str] = None
    ministry: Optional[str] = None
    tender_type: Optional[str] = None
    close_date: Optional[str] = None
    pdf_url: Optional[str] = None


@dataclass
class TenderDetails:
    tender_no: Optional[str] = None
    tender_type: Optional[str] = None
    subject: Optional[str] = None
    ministry: Optional[str] = None
    entity_tender_no: Optional[str] = None
    request_types: Optional[str] = None
    envelopes_system: Optional[str] = None
    tender_bond: Optional[str] = None
    documents_value_qr: Optional[str] = None
    closing_date: Optional[str] = None

    brief_description: Optional[str] = None
    targeted_tenderer_type: Optional[str] = None
    service_delivery_method: Optional[str] = None
    auction_type: Optional[str] = None
    local_value_system: Optional[str] = None
    tender_validity_period: Optional[str] = None
    evaluation_basis: Optional[str] = None

    activities: List[Dict[str, str]] = None
    special_conditions: Dict[str, str] = None
    general_conditions_text: Optional[str] = None


# -----------------------------
# Utilities
# -----------------------------
def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def safe_get_text(el) -> str:
    return clean_text(el.get_text(" ", strip=True)) if el else ""


def extract_tender_id(details_url: str) -> Optional[str]:
    m = re.search(r"/TenderDetails/(\d+)", details_url)
    return m.group(1) if m else None


def summarize_topic(subject: str, brief_desc: str, max_len: int = 160) -> str:
    combined = clean_text(" ".join([subject or "", brief_desc or ""]))
    combined = re.sub(r"\b[A-Z]{2,}/[A-Z0-9\-]{2,}/[A-Z0-9\-]{2,}\b", "", combined)
    combined = clean_text(combined)
    if not combined:
        return ""
    if len(combined) <= max_len:
        return combined
    cut = combined[:max_len]
    last = max(cut.rfind("."), cut.rfind("،"), cut.rfind(","), cut.rfind(";"), cut.rfind(":"))
    if last > max_len * 0.65:
        cut = cut[:last]
    return clean_text(cut.rstrip(" ,.;:-")) + "…"


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def parse_date_any(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    t = clean_text(s)
    t = t.replace("\\", "/").replace("-", "/").replace(".", "/")

    fmts = ["%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y", "%d/%m/%Y %H:%M", "%Y/%m/%d %H:%M", "%d/%m/%Y %I:%M %p"]
    for f in fmts:
        try:
            return datetime.strptime(t, f).date()
        except Exception:
            pass

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", t)
    if m:
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mth, d)
        except Exception:
            return None

    m2 = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", t)
    if m2:
        y, mth, d = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        try:
            return date(y, mth, d)
        except Exception:
            return None

    return None


def close_date_key(item: Dict[str, Any]) -> date:
    """Safe sort key for closing date; pushes missing dates to far future."""
    landing = item.get("landing") or {}
    close_s = landing.get("close_date")
    d = parse_date_any(close_s)
    return d if d else date(9999, 12, 31)


def read_last_successful_date() -> date:
    if LAST_SCRAPE_FILE.exists():
        try:
            txt = LAST_SCRAPE_FILE.read_text(encoding="utf-8").strip()
            return datetime.strptime(txt, "%Y-%m-%d").date()
        except Exception:
            pass
    # default: last 7 days if none stored
    return date.today() - timedelta(days=7)


def write_last_successful_date(d: date) -> None:
    try:
        LAST_SCRAPE_FILE.write_text(d.strftime("%Y-%m-%d"), encoding="utf-8")
    except Exception:
        pass


def load_default_interests_text() -> str:
    if DEFAULT_INTERESTS_PATH.exists():
        try:
            return DEFAULT_INTERESTS_PATH.read_text(encoding="utf-8")
        except Exception:
            pass
    return json.dumps(DEFAULT_INTERESTS_FALLBACK, ensure_ascii=False, indent=2)


def load_default_isic_text() -> str:
    if ISIC_DEFAULT_PATH.exists():
        return ISIC_DEFAULT_PATH.read_text(encoding="utf-8")
    # fallback to empty list
    return "[]"


# -----------------------------
# ISIC config parsing + matching
# -----------------------------
def _norm_code(x: Any) -> str:
    """Keep ISIC code as text; preserve leading zeros when present in source."""
    if x is None:
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+", "", s)
    return s


def _norm_txt(x: Any) -> str:
    return clean_text(str(x or "")).lower()


def _norm_key(k: str) -> str:
    """Normalize JSON keys for tolerant matching (handles Arabic diacritics + whitespace/punctuation)."""
    if k is None:
        return ""
    s = str(k)
    s = s.replace("\ufeff", "").replace("\u200f", "").replace("\u200e", "")
    s = s.strip().lower()
    # remove Arabic diacritics + tatweel
    s = re.sub(r"[\u064B-\u065F\u0670\u06D6-\u06ED\u0640]", "", s)
    # normalize unicode
    s = unicodedata.normalize("NFKC", s)
    # collapse spaces
    s = re.sub(r"\s+", " ", s)
    # remove common punctuation
    s = re.sub(r"[:\-_/\\\(\)\[\]\{\}\|]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _get_any(row: Dict[str, Any], candidates: List[str]) -> Any:
    """Get a value from dict row by trying multiple candidate keys, with normalized matching."""
    if not isinstance(row, dict):
        return None
    # direct attempts first
    for c in candidates:
        if c in row:
            return row.get(c)
    # normalized map
    norm_map = {_norm_key(k): k for k in row.keys()}
    for c in candidates:
        nk = _norm_key(c)
        if nk in norm_map:
            return row.get(norm_map[nk])
    return None


def parse_isic_config(isic_text: str) -> List[Dict[str, str]]:
    """
    Parse ISIC list from JSON array of objects.

    This is intentionally tolerant because teams often generate JSON from:
    - Excel exports
    - manual conversions
    - mixed Arabic/English headers
    - duplicated header rows inside the data

    Returns: list of dicts with canonical keys:
      - activity_code (TEXT; keeps leading zeros)
      - activity_name (Arabic or general)
      - activity_name_english
    """
    try:
        data = json.loads(isic_text) if (isic_text or "").strip() else []
    except Exception as e:
        raise ValueError(f"Invalid ISIC JSON: {e}")

    if data is None:
        data = []
    if isinstance(data, dict) and isinstance(data.get("items"), list):
        data = data["items"]
    if not isinstance(data, list):
        raise ValueError("ISIC list must be a JSON array (list) of objects.")

    code_candidates = [
        "activity_code", "Activity code", "Activity Code", "code", "Code", "ISIC Code", "isic_code", "ISIC",
        "Activity", "النشاط", "رمز النشاط", "رمز_النشاط", "كود النشاط", "كود_النشاط", "رقم النشاط", "رقم_النشاط",
    ]
    ar_candidates = [
        "activity_name", "Activity name", "Activity Name", "name", "Name", "Activity Name (AR)",
        "اسم النشاط", "اسم_النشاط", "اسم نشاط", "اسم_نشاط",
    ]
    en_candidates = [
        "activity_name_english", "Activity Name English", "Activity name English", "Activity Name (EN)",
        "name_english", "english_name", "ActivityNameEnglish",
        "اسم النشاط انجليزي", "اسم النشاط إنجليزي", "اسم النشاط الانجليزي", "اسم النشاط بالانجليزي",
        "اسم_النشاط_انجليزي", "اسم_النشاط_بالانجليزي",
    ]

    # header-like values to skip
    header_code_vals = {"activity", "activitycode", "isic", "isiccode", "code", "النشاط", "رمزالنشاط"}
    header_name_vals = {"activityname", "activitynameenglish", "اسمالنشاط", "اسمالنشاطانجليزي", "اسمالنشاطالانجليزي"}

    out: List[Dict[str, str]] = []
    seen = set()

    for row in data:
        if not isinstance(row, dict):
            continue

        code = _get_any(row, code_candidates)
        name_ar = _get_any(row, ar_candidates)
        name_en = _get_any(row, en_candidates)

        # Fallback: search keys by normalized patterns if still missing
        if name_ar is None and name_en is None:
            for k, v in row.items():
                nk = _norm_key(k)
                if "activityname" in nk or "اسمالنشاط" in nk:
                    name_ar = v
                    break

        if name_en is None:
            for k, v in row.items():
                nk = _norm_key(k)
                if "english" in nk or "انجليزي" in nk or "بالانجليزي" in nk:
                    name_en = v
                    break

        code_s = _norm_code(code)
        name_ar_s = clean_text(name_ar or "")
        name_en_s = clean_text(name_en or "")

        # Skip header rows mistakenly included in data
        if _norm_key(code_s) in header_code_vals:
            continue
        if _norm_key(name_ar_s) in header_name_vals or _norm_key(name_en_s) in header_name_vals:
            continue

        if not (code_s or name_ar_s or name_en_s):
            continue

        key = (code_s, name_ar_s, name_en_s)
        if key in seen:
            continue
        seen.add(key)

        out.append({
            "activity_code": code_s,
            "activity_name": name_ar_s,
            "activity_name_english": name_en_s,
        })

    return out


def match_isic_activities(
    tender_activities: List[Dict[str, str]],
    isic_cfg: List[Dict[str, str]],
) -> Dict[str, Any]:
    """
    Offline deterministic matching:
      - exact code match
      - conservative token-overlap name match as fallback
    """
    cfg_by_code = {_norm_code(x.get("activity_code")): x for x in isic_cfg if _norm_code(x.get("activity_code"))}

    cfg_names = []
    for x in isic_cfg:
        ar = _norm_txt(x.get("activity_name"))
        en = _norm_txt(x.get("activity_name_english"))
        if ar:
            cfg_names.append((ar, x, "ar"))
        if en:
            cfg_names.append((en, x, "en"))

    matched_items = []
    seen = set()

    for a in (tender_activities or []):
        tcode = _norm_code(a.get("activity_code") or a.get("code"))
        tname = _norm_txt(a.get("activity_name") or a.get("name"))

        # 1) code
        if tcode and tcode in cfg_by_code and tcode not in seen:
            item = cfg_by_code[tcode]
            matched_items.append({
                "activity_code": tcode,
                "activity_name": item.get("activity_name") or item.get("activity_name_english") or "",
                "reason": "code"
            })
            seen.add(tcode)
            continue

        # 2) name fallback
        if not tname:
            continue

        t_tokens = {w for w in re.split(r"[^\w\u0600-\u06FF]+", tname) if len(w) >= 3}
        if not t_tokens:
            continue

        best_item = None
        best_code = ""
        best_score = 0.0
        best_lang = ""

        for cname, item, lang in cfg_names:
            c_tokens = {w for w in re.split(r"[^\w\u0600-\u06FF]+", cname) if len(w) >= 3}
            if not c_tokens:
                continue
            inter = len(t_tokens & c_tokens)
            union = len(t_tokens | c_tokens)
            score = inter / union if union else 0.0
            if score > best_score:
                best_score = score
                best_item = item
                best_code = _norm_code(item.get("activity_code"))
                best_lang = lang

        # conservative threshold
        if best_item is not None and best_score >= 0.45 and best_code and best_code not in seen:
            matched_items.append({
                "activity_code": best_code,
                "activity_name": best_item.get("activity_name") or best_item.get("activity_name_english") or "",
                "reason": f"name_{best_lang}:{best_score:.2f}"
            })
            seen.add(best_code)

    return {
        "match_count": len(matched_items),
        "matched_codes": [m["activity_code"] for m in matched_items],
        "matched_items": matched_items,
    }


# -----------------------------
# HTML parsing (landing)
# -----------------------------
def parse_landing_html(html: str) -> List[TenderCard]:
    soup = BeautifulSoup(html, "html.parser")
    tenders: List[TenderCard] = []

    for row in soup.select("div.row.custom-cards"):
        tender_no = clean_text(
            row.select_one("div.col-md-7.cards-col .col-header span.card-label").get_text(" ", strip=True)
        ) if row.select_one("div.col-md-7.cards-col .col-header span.card-label") else ""

        a = row.select_one("div.col-md-7.cards-col .col-header span.card-title a")
        title = safe_get_text(a)
        details_url = a.get("href", "").strip() if a else ""
        details_url = urljoin(BASE_URL, details_url)

        if not details_url or "/TendersOnlineServices/TenderDetails/" not in details_url:
            continue

        tender_id = extract_tender_id(details_url)

        label_value = {}
        for r in row.select("div.col-md-7.cards-col .col-footer div.cards-row"):
            label = safe_get_text(r.select_one("span.card-label"))
            value = safe_get_text(r.select_one("span.card-title"))
            if label and value:
                label_value[label.lower()] = value

        ministry = safe_get_text(row.select_one("div.col-md-3.cards-col .col-header span.card-title span"))
        tender_type = safe_get_text(row.select_one("div.col-md-3.cards-col .col-footer div.cards-row span.card-title span"))

        close_date = ""
        close_span = row.select_one("div.col-md-2.circle-container span.card-label span:nth-of-type(2)")
        if close_span:
            close_date = safe_get_text(close_span)

        pdf_url = None
        pdf_a = row.select_one("a[href*='/Main/GetTenderDetailsFile/'][href*='type=2']")
        if pdf_a and pdf_a.get("href"):
            pdf_url = urljoin(BASE_URL, pdf_a["href"].strip())

        tenders.append(
            TenderCard(
                tender_no=tender_no,
                title=title,
                details_url=details_url,
                tender_id=tender_id,
                publish_date=label_value.get("publish date"),
                requested_sector_type=label_value.get("requested sector type"),
                tender_bond_qar=label_value.get("tender bond (qar)"),
                documents_value_qar=label_value.get("documents value (qr)"),
                ministry=ministry or None,
                tender_type=tender_type or None,
                close_date=close_date or None,
                pdf_url=pdf_url,
            )
        )

    return tenders


# -----------------------------
# HTML parsing (details) - position-based
# -----------------------------
def parse_details_html(html: str) -> TenderDetails:
    """
    Parse Tender Details HTML into a TenderDetails object.

    IMPORTANT:
    - "Activities list" table is not always found via an exact <h2 string="..."> match
      (sometimes the <h2> contains nested tags/whitespace).
    - Some pages may not have the "Activities list" heading at all, but still have a table
      with headers like "Activity code" / "Activity name" (or Arabic equivalents).
    So we detect the activities table by its column headers first, then fall back to heading-based lookup.
    """
    soup = BeautifulSoup(html, "html.parser")
    details = TenderDetails(activities=[], special_conditions={})
    tables = soup.select("table.custom--table")

    # --- Main details table (first custom--table) ---
    if len(tables) >= 1:
        row = tables[0].select_one("tbody tr")
        if row:
            cells = [clean_text(td.get_text(" ", strip=True)) for td in row.select("td")]
            if len(cells) >= 10:
                details.tender_no = cells[0] or None
                details.tender_type = cells[1] or None
                details.subject = cells[2] or None
                details.ministry = cells[3] or None
                details.entity_tender_no = cells[4] or None
                details.request_types = cells[5] or None
                details.envelopes_system = cells[6] or None
                details.tender_bond = cells[7] or None
                details.documents_value_qr = cells[8] or None
                details.closing_date = cells[9] or None

    # --- Brief description / targeted tenderer / etc (second custom--table if present) ---
    if len(tables) >= 2:
        row2 = tables[1].select_one("tbody tr")
        if row2:
            cells = [clean_text(td.get_text(" ", strip=True)) for td in row2.select("td")]
            if len(cells) >= 7:
                details.brief_description = cells[0] or None
                details.targeted_tenderer_type = cells[1] or None
                details.service_delivery_method = cells[2] or None
                details.auction_type = cells[3] or None
                details.local_value_system = cells[4] or None
                details.tender_validity_period = cells[5] or None
                details.evaluation_basis = cells[6] or None

    # -------------------------
    # Activities list extraction
    # -------------------------
    def _is_activities_table(table) -> bool:
        headers = [clean_text(th.get_text(" ", strip=True)).lower() for th in table.select("thead th")]
        if not headers:
            return False

        # English / Arabic header checks
        has_code = any(("activity code" in h) or ("isic" in h and "code" in h) or ("رمز" in h and "نشاط" in h) for h in headers)
        has_name = any(("activity name" in h) or ("اسم" in h and "نشاط" in h) for h in headers)

        # Some pages might show "Activity name English" too — still OK.
        return has_code and has_name

    activities_table = None

    # 1) Prefer header-based detection across ALL tables
    for tbl in soup.select("table"):
        if _is_activities_table(tbl):
            activities_table = tbl
            break

    # 2) Fallback: find heading by text (robust: use get_text, not `.string`)
    if activities_table is None:
        for tag in soup.find_all(["h2", "h3", "h4"]):
            t = clean_text(tag.get_text(" ", strip=True)).lower()
            if ("activities list" in t) or ("قائمة الأنشطة" in t) or ("قائمة النشاط" in t):
                activities_table = tag.find_next("table")
                break

    # 3) Parse rows
    if activities_table is not None:
        for tr in activities_table.select("tbody tr"):
            # Sometimes rows might contain <th> cells; allow both.
            cells = tr.find_all(["td", "th"])
            if len(cells) >= 2:
                code = clean_text(cells[0].get_text(" ", strip=True))
                name = clean_text(cells[1].get_text(" ", strip=True))
                if code or name:
                    # Keep code as TEXT (preserve leading zeros)
                    details.activities.append({
                        "activity_code": str(code) if code is not None else "",
                        "activity_name": name or ""
                    })

    # -------------------------
    # Special conditions extraction
    # -------------------------
    def _find_heading(text_pat: str):
        rx = re.compile(text_pat, re.I)
        # Robust: match on get_text (not .string)
        for tag in soup.find_all(["h2", "h3", "h4"]):
            if rx.search(clean_text(tag.get_text(" ", strip=True))):
                return tag
        return None

    h_sc = _find_heading(r"Special Conditions|الشروط الخاصة")
    if h_sc:
        sc_table = h_sc.find_next("table")
        if sc_table:
            for tr in sc_table.select("tbody tr"):
                th = tr.select_one("th")
                td = tr.select_one("td")
                if th and td:
                    details.special_conditions[clean_text(th.get_text(" ", strip=True))] = clean_text(td.get_text(" ", strip=True))

    h_gc = _find_heading(r"General Conditions|الشروط العامة")
    if h_gc:
        gc_div = h_gc.find_next("div")
        if gc_div:
            # keep full text
            details.general_conditions_text = clean_text(gc_div.get_text(" ", strip=True)) or None

    return details


# -----------------------------
# Fetching (Playwright async)
# -----------------------------
async def fetch_html_playwright(url: str, timeout_ms: int = 60000) -> str:
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
        )
        page = await context.new_page()

        await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)

        # ✅ wait for tender items to appear (very important on Streamlit Cloud)
        try:
            await page.wait_for_selector('a[href*="TenderDetails"]', timeout=timeout_ms)
        except Exception:
            # keep going (we’ll return whatever HTML we got)
            pass

        html = await page.content()
        await context.close()
        await browser.close()
        return html


async def fetch_many_details(urls: List[str], concurrency: int = 6) -> Dict[str, str]:
    sem = asyncio.Semaphore(concurrency)
    results: Dict[str, str] = {}

    async def worker(u: str):
        async with sem:
            try:
                results[u] = await fetch_html_playwright(u)
            except Exception as e:
                results[u] = f"__ERROR__ {e}"

    await asyncio.gather(*(worker(u) for u in urls))
    return results


def run_async(coro):
    # robust sync runner for Streamlit
    if os.name == "nt":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(coro)
    finally:
        try:
            loop.close()
        except Exception:
            pass


# -----------------------------
# Embeddings + scoring (stand-out based)
# -----------------------------
def cosine(a: List[float], b: List[float]) -> float:
    dot = 0.0
    na = 0.0
    nb = 0.0
    for x, y in zip(a, b):
        dot += x * y
        na += x * x
        nb += y * y
    if na == 0.0 or nb == 0.0:
        return 0.0
    return dot / ((na ** 0.5) * (nb ** 0.5))


@st.cache_resource
def get_embedder():
    cache_dir = os.path.join(os.path.expanduser("~"), ".cache", "fastembed_models")
    return TextEmbedding(model_name="intfloat/multilingual-e5-large", cache_dir=cache_dir)


@st.cache_data
def build_interest_vectors_cached(interests_json_text: str) -> Dict[str, List[float]]:
    interests = json.loads(interests_json_text)
    embedder = get_embedder()

    names = list(interests.keys())
    texts = []
    for n in names:
        parts = [p for p in (interests.get(n) or []) if (p or "").strip()]
        doc = n + " :: " + " | ".join(parts[:20])
        texts.append("query: " + doc)

    vecs = [list(v) for v in embedder.embed(texts)]
    return {n: v for n, v in zip(names, vecs)}


def semantic_relevance_standout(
    interest_vecs: Dict[str, List[float]],
    tender_text: str,
    *,
    top_k: int = 8,
    margin_gate: float = 0.018,
    z_gate: float = 0.75,
):
    embedder = get_embedder()
    tvec = list(next(embedder.embed(["passage: " + tender_text])))

    sims = [(name, cosine(tvec, ivec)) for name, ivec in interest_vecs.items()]
    sims.sort(key=lambda x: x[1], reverse=True)

    best_name, best_sim = sims[0]
    second_sim = sims[1][1] if len(sims) > 1 else 0.0
    margin = best_sim - second_sim

    values = [s for _, s in sims]
    mean = sum(values) / len(values)
    var = sum((x - mean) ** 2 for x in values) / max(1, (len(values) - 1))
    std = math.sqrt(var)
    z = (best_sim - mean) / (std + 1e-9)

    raw = 1.35 * z + 22.0 * margin - 1.55
    conf = 1.0 / (1.0 + math.exp(-raw))

    if margin < margin_gate or z < z_gate:
        score = 0
        label = "Irrelevant / Unclear"
    else:
        score = int(round(conf * 100))
        if score >= 70:
            label = "Relevant (High)"
        elif score >= 40:
            label = "Relevant (Medium)"
        else:
            label = "Relevant (Low)"

    breakdown = [{"interest": n, "sim": round(s, 4)} for n, s in sims[:top_k]]
    debug = {
        "best_interest": best_name,
        "best_sim": round(best_sim, 4),
        "second_sim": round(second_sim, 4),
        "margin": round(margin, 4),
        "z": round(z, 4),
        "confidence": round(conf, 4),
        "label": label,
        "gates": {"margin_gate": margin_gate, "z_gate": z_gate},
    }
    explanation = (
        f"Best={best_name} (sim={best_sim:.4f}), margin={margin:.4f}, z={z:.2f}, conf={conf:.2f} => {label}"
    )
    return score, breakdown, debug, explanation


# -----------------------------
# Export helpers (CSV/Excel)
# -----------------------------
def flatten_row(item: Dict[str, Any]) -> Dict[str, str]:
    landing = item.get("landing") or {}
    details = item.get("details") or {}
    dbg = item.get("relevance_debug") or {}

    # first columns requested
    row = {
        "tender_no": str(landing.get("tender_no", "")),
        "title": str(landing.get("title", "")),
        "close_date": str(landing.get("close_date", "")),
    }

    # rest
    row.update({
        "relevance_score": str(item.get("relevance_score", "")),
        "label": str(dbg.get("label", "")),
        "overall_relevance_reason": str(item.get("overall_relevance_reason", "")),
        "best_interest": str(dbg.get("best_interest", "")),
        "publish_date": str(landing.get("publish_date", "")),
        "ministry": str(landing.get("ministry", "")),
        "tender_type": str(landing.get("tender_type", "")),
        "topic_summary": str(item.get("topic_summary", "")),
        "margin": str(dbg.get("margin", "")),
        "z": str(dbg.get("z", "")),
        "confidence": str(dbg.get("confidence", "")),
        "best_sim": str(dbg.get("best_sim", "")),
        "second_sim": str(dbg.get("second_sim", "")),
        "details_url": str(landing.get("details_url", "")),
        "pdf_url": str(landing.get("pdf_url", "")),
        "details_subject": str(details.get("subject", "")) if details else "",
        "details_brief": str(details.get("brief_description", "")) if details else "",
        "explanation": str(item.get("relevance_explanation", "")),
        "isic_match_count": str(item.get("isic_match_count", "")),
        "isic_matched_codes": "|".join([str(x) for x in (item.get("isic_matched_codes") or [])]),
        "overall_relevant": str(item.get("overall_relevant", "")),
        "scraped_at_utc": str(item.get("scraped_at_utc", "")),
    })
    return row


def build_html_table(items: List[Dict[str, Any]], title: str) -> str:
    items = sorted(items, key=close_date_key)  # sort by closing date asc

    def esc(x: str) -> str:
        return (x or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def badge(label: str) -> Tuple[str, str]:
        # returns (bg, fg)
        l = (label or "").lower()
        if "high" in l:
            return ("#0B5FFF", "white")
        if "medium" in l:
            return ("#7A5AF8", "white")
        if "low" in l:
            return ("#6B7280", "white")
        if "irrelevant" in l or "unclear" in l:
            return ("#111827", "white")
        return ("#374151", "white")

    def reason_badge(reason: str) -> Tuple[str, str]:
        r = (reason or "").lower()
        if "embedding" in r and "isic" in r:
            return ("rgba(14,165,233,0.25)", "rgba(255,255,255,0.92)")  # sky
        if "embedding" in r:
            return ("rgba(99,102,241,0.25)", "rgba(255,255,255,0.92)")  # indigo
        if "isic" in r:
            return ("rgba(16,185,129,0.25)", "rgba(255,255,255,0.92)")  # green
        return ("rgba(255,255,255,0.10)", "rgba(255,255,255,0.88)")

    def score_bar(score_int: int) -> str:
        s = max(0, min(100, int(score_int)))
        if s >= 70:
            c = "#0B5FFF"
        elif s >= 40:
            c = "#F59E0B"
        else:
            c = "#9CA3AF"
        return f"""
        <div class="score-wrap">
          <div class="score-bar" style="width:{s}%; background:{c};"></div>
        </div>
        """

    rows_html = []
    for it in items:
        landing = it.get("landing") or {}
        dbg = it.get("relevance_debug") or {}

        details_url = landing.get("details_url", "")
        tender_no = esc(str(landing.get("tender_no", "")))
        title_txt = esc(str(landing.get("title", "")))
        close_dt = esc(str(landing.get("close_date", "")))
        ministry = esc(str(landing.get("ministry", "")))

        score = int(it.get("relevance_score") or 0)
        label = esc(str(dbg.get("label", "")))
        best_interest = esc(str(dbg.get("best_interest", "")))
        reason = esc(str(it.get("overall_relevance_reason", "")))

        bg, fg = badge(label)
        rbg, rfg = reason_badge(reason)

        tender_no_link = (
            f'<a class="tlink" href="{esc(details_url)}" target="_blank" rel="noopener noreferrer">{tender_no}</a>'
            if details_url else tender_no
        )

        rows_html.append(
            f"""
            <tr>
              <td class="col-no">{tender_no_link}</td>
              <td class="col-title" title="{title_txt}">{title_txt}</td>
              <td class="col-close">{close_dt}</td>
              <td class="col-score">
                <div class="score-cell">
                  <div class="score-num">{score}</div>
                  {score_bar(score)}
                </div>
              </td>
              <td class="col-label">
                <span class="pill" style="background:{bg}; color:{fg};">{label}</span>
              </td>
              <td class="col-reason">
                <span class="pill pill-reason" style="background:{rbg}; color:{rfg};">{reason}</span>
              </td>
              <td class="col-interest">{best_interest}</td>
              <td class="col-ministry">{ministry}</td>
            </tr>
            """
        )

    return f"""
    <style>
      :root {{
        --border: rgba(255,255,255,0.08);
        --border2: rgba(255,255,255,0.12);
        --bg: rgba(255,255,255,0.04);
        --bg2: rgba(255,255,255,0.06);
        --text: rgba(255,255,255,0.92);
        --muted: rgba(255,255,255,0.72);
        --shadow: 0 10px 25px rgba(0,0,0,0.25);
      }}

      .card {{
        border: 1px solid var(--border);
        background: rgba(17, 24, 39, 0.35);
        border-radius: 16px;
        padding: 14px 14px 10px;
        box-shadow: var(--shadow);
      }}
      .title {{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap: 12px;
        margin: 0 0 10px 0;
        font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
      }}
      .title h3 {{
        margin:0;
        font-size: 16px;
        font-weight: 650;
        color: var(--text);
      }}

      .tw {{
        width: 100%;
        overflow: auto;
        border-radius: 14px;
        border: 1px solid var(--border2);
      }}

      table {{
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
        font-size: 13px;
        color: var(--text);
        background: rgba(0,0,0,0.15);
      }}

      thead th {{
        position: sticky;
        top: 0;
        z-index: 2;
        text-align: left;
        padding: 10px 10px;
        font-size: 12px;
        letter-spacing: .02em;
        text-transform: uppercase;
        color: var(--muted);
        background: rgba(17, 24, 39, 0.92);
        border-bottom: 1px solid var(--border2);
        white-space: nowrap;
      }}

      tbody td {{
        padding: 10px 10px;
        border-bottom: 1px solid var(--border);
        vertical-align: middle;
      }}

      tbody tr:nth-child(odd) {{
        background: rgba(255,255,255,0.03);
      }}
      tbody tr:hover {{
        background: rgba(11, 95, 255, 0.10);
      }}

      .col-no {{ white-space: nowrap; width: 120px; }}
      .col-close {{ white-space: nowrap; width: 120px; }}
      .col-score {{ white-space: nowrap; width: 140px; }}
      .col-label {{ white-space: nowrap; width: 150px; }}
      .col-reason {{ white-space: nowrap; width: 160px; }}
      .col-interest {{ width: 220px; }}
      .col-ministry {{ width: 220px; }}
      .col-title {{
        max-width: 540px;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
      }}

      a.tlink {{
        color: #93C5FD;
        text-decoration: none;
        font-weight: 650;
      }}
      a.tlink:hover {{
        text-decoration: underline;
      }}

      .pill {{
        display:inline-flex;
        align-items:center;
        border-radius: 999px;
        padding: 5px 9px;
        font-size: 12px;
        font-weight: 650;
        line-height: 1;
      }}
      .pill-reason {{
        border: 1px solid rgba(255,255,255,0.12);
      }}

      .score-cell {{
        display:flex;
        align-items:center;
        gap: 10px;
      }}
      .score-num {{
        width: 32px;
        text-align: right;
        font-weight: 700;
        color: var(--text);
      }}
      .score-wrap {{
        flex: 1;
        height: 8px;
        background: rgba(255,255,255,0.12);
        border-radius: 999px;
        overflow: hidden;
        border: 1px solid rgba(255,255,255,0.10);
      }}
      .score-bar {{
        height: 100%;
        border-radius: 999px;
      }}
    </style>

    <div class="card">
      <div class="title">
        <h3>{esc(title)}</h3>
      </div>
      <div class="tw">
        <table>
          <thead>
            <tr>
              <th>Tender No</th>
              <th>Title</th>
              <th>Closing</th>
              <th>Score</th>
              <th>Label</th>
              <th>Reason</th>
              <th>Best Interest</th>
              <th>Ministry</th>
            </tr>
          </thead>
          <tbody>
            {''.join(rows_html) if rows_html else '<tr><td colspan="8" style="padding:14px;color:rgba(255,255,255,0.75)">No rows</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """


def to_csv_bytes(items: List[Dict[str, Any]]) -> bytes:
    if not items:
        return b""
    rows = [flatten_row(x) for x in items]
    header = list(rows[0].keys())

    from io import StringIO
    buff = StringIO()
    writer = csv.DictWriter(buff, fieldnames=header)
    writer.writeheader()
    writer.writerows(rows)
    return buff.getvalue().encode("utf-8")


def to_excel_bytes(items: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    rows = [flatten_row(x) for x in items]
    if not rows:
        ws.append(["No data"])
    else:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])

        # auto width (simple)
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(h)
            for rr in rows[:200]:
                max_len = max(max_len, len(str(rr.get(h, ""))))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))

        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# Pipeline (async)
# -----------------------------
async def run_pipeline_pages(
    start_page: int,
    max_pages: int,
    max_tenders_total: int,
    concurrency: int,
) -> List[Dict[str, Any]]:

    all_cards: List[TenderCard] = []
    seen_keys: set[str] = set()

    empty_streak = 0

    for p in range(start_page, start_page + max_pages):
        landing_url = LANDING_URL_TEMPLATE.format(page=p)

        landing_html = await fetch_html_playwright(landing_url, timeout_ms=60000)
        cards = parse_landing_html(landing_html)

        # retry once with longer timeout
        if not cards:
            landing_html = await fetch_html_playwright(landing_url, timeout_ms=120000)
            cards = parse_landing_html(landing_html)

        if not cards:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        else:
            empty_streak = 0

        # ✅ DEDUP per-page + across pages (by details_url primarily)
        for c in cards:
            # Prefer details_url as the true unique key
            key = (getattr(c, "details_url", None) or "").strip()

            # Fallback key if details_url missing for any reason
            if not key:
                tn = (getattr(c, "tender_no", None) or "").strip()
                title = (getattr(c, "tender_name", None) or getattr(c, "name", None) or getattr(c, "title", None) or "").strip()
                key = f"{tn}||{title}".strip()

            # If still empty, skip (should not happen, but safe)
            if not key:
                continue

            if key in seen_keys:
                continue

            seen_keys.add(key)
            all_cards.append(c)

        if max_tenders_total > 0 and len(all_cards) >= max_tenders_total:
            all_cards = all_cards[:max_tenders_total]
            break

    # ✅ Safety: final dedup again (keeps order)
    uniq_cards: List[TenderCard] = []
    seen2: set[str] = set()
    for c in all_cards:
        key = (getattr(c, "details_url", None) or "").strip()
        if not key:
            tn = (getattr(c, "tender_no", None) or "").strip()
            title = (getattr(c, "tender_name", None) or getattr(c, "name", None) or getattr(c, "title", None) or "").strip()
            key = f"{tn}||{title}".strip()
        if not key or key in seen2:
            continue
        seen2.add(key)
        uniq_cards.append(c)

    all_cards = uniq_cards

    # Fetch details for unique URLs only
    details_html_map = await fetch_many_details(
        [c.details_url for c in all_cards if getattr(c, "details_url", None)],
        concurrency=concurrency
    )

    results: List[Dict[str, Any]] = []
    for c in all_cards:
        raw = details_html_map.get(c.details_url, "") if getattr(c, "details_url", None) else ""
        details = None
        if raw and not raw.startswith("__ERROR__"):
            try:
                details = parse_details_html(raw)
            except Exception:
                details = None

        results.append({
            "landing": asdict(c),
            "details": asdict(details) if details else None,
        })

    return results



# -----------------------------
# Editors (Table-based) for Interests + ISIC (popup dialogs)
# -----------------------------
def interests_to_rows(interests_obj: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    if not isinstance(interests_obj, dict):
        return rows
    for k, v in interests_obj.items():
        name = clean_text(str(k))
        if not name:
            continue
        if isinstance(v, list):
            for line in v:
                txt = clean_text(str(line))
                if txt:
                    rows.append({"interest": name, "text": txt})
        elif isinstance(v, str):
            txt = clean_text(v)
            if txt:
                rows.append({"interest": name, "text": txt})
        else:
            # ignore unsupported value types
            continue
    if not rows:
        # keep at least one editable row
        rows = [{"interest": "", "text": ""}]
    return rows


def rows_to_interests(rows: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        name = clean_text(str(r.get("interest", "")))
        txt = clean_text(str(r.get("text", "")))
        if not name or not txt:
            continue
        out.setdefault(name, []).append(txt)
    return out


def isic_cfg_to_rows(isic_cfg: List[Dict[str, str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for r in isic_cfg or []:
        if not isinstance(r, dict):
            continue
        rows.append({
            "activity_code": _norm_code(r.get("activity_code")),
            "activity_name": clean_text(r.get("activity_name") or ""),
            "activity_name_english": clean_text(r.get("activity_name_english") or ""),
        })
    if not rows:
        rows = [{"activity_code": "", "activity_name": "", "activity_name_english": ""}]
    return rows


def rows_to_isic_cfg(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    seen = set()
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        code = _norm_code(r.get("activity_code"))
        name = clean_text(r.get("activity_name") or "")
        en = clean_text(r.get("activity_name_english") or "")
        if not (code or name or en):
            continue
        # preserve leading zeros: keep as string
        key = (code, name, en)
        if key in seen:
            continue
        seen.add(key)
        out.append({"activity_code": code, "activity_name": name, "activity_name_english": en})
    return out


def _save_text_file(path: Path, text: str) -> Tuple[bool, str]:
    try:
        path.write_text(text, encoding="utf-8")
        return True, "Saved."
    except Exception as e:
        return False, str(e)


# -----------------------------
# Editors UI (popover-based) for Interests + ISIC (table editing, no dialogs)
# -----------------------------
HAS_POPOVER = hasattr(st, "popover")


def render_interests_editor_ui():
    st.caption("Edit interests as a table. Each row = one text line under an interest.")

    # Parse current JSON
    try:
        interests_obj = json.loads(st.session_state.get("interests_text", "{}") or "{}")
        if not isinstance(interests_obj, dict):
            interests_obj = {}
    except Exception:
        interests_obj = {}

    rows = interests_to_rows(interests_obj)

    edited = st.data_editor(
        rows,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "interest": st.column_config.TextColumn("Interest", required=False, width="medium"),
            "text": st.column_config.TextColumn("Description / Keywords", required=False, width="large"),
        },
        key="_interests_editor_table",
    )

    c1, c2 = st.columns([1, 2])
    with c1:
        if st.button("✅ Save interests.json", type="primary", key="_save_interests_btn"):
            new_obj = rows_to_interests(edited)
            if not new_obj:
                st.error("Interests cannot be empty. Add at least one row with Interest + Description.")
                return
            new_text = json.dumps(new_obj, ensure_ascii=False, indent=2)
            st.session_state["interests_text"] = new_text
            ok, msg = _save_text_file(DEFAULT_INTERESTS_PATH, new_text)
            if ok:
                st.success("Saved to interests.json")
            else:
                st.error(f"Could not save file: {msg}")
    with c2:
        st.caption("Tip: keep each description line short and specific (EN/AR allowed).")


def render_isic_editor_ui():
    st.caption("Edit ISIC-of-interests as a table. Activity code is treated as TEXT (leading zeros preserved).")

    # Parse current
    try:
        cfg = parse_isic_config(st.session_state.get("isic_text", "[]") or "[]")
    except Exception:
        cfg = []

    rows = isic_cfg_to_rows(cfg)

    edited = st.data_editor(
        rows,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "activity_code": st.column_config.TextColumn("Activity code", required=False, width="small"),
            "activity_name": st.column_config.TextColumn("Activity name (AR)", required=False, width="large"),
            "activity_name_english": st.column_config.TextColumn("Activity name (EN)", required=False, width="large"),
        },
        key="_isic_editor_table",
    )

    c1, c2 = st.columns([1, 2])
    with c1:
        if st.button("✅ Save isic-of-interests.json", type="primary", key="_save_isic_btn"):
            cfg2 = rows_to_isic_cfg(edited)
            new_text = json.dumps(cfg2, ensure_ascii=False, indent=2)
            st.session_state["isic_text"] = new_text
            ok, msg = _save_text_file(ISIC_DEFAULT_PATH, new_text)
            if ok:
                st.success("Saved to isic-of-interests.json")
            else:
                st.error(f"Could not save file: {msg}")
    with c2:
        st.caption("Paste from Excel by copying columns into the table rows.")


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="Monaqasat Tender Radar", layout="wide")
st.title("Monaqasat Tender Radar (Offline AI Matching)")

# Load defaults once
if "interests_text" not in st.session_state:
    st.session_state["interests_text"] = load_default_interests_text()

if "after_date" not in st.session_state:
    st.session_state["after_date"] = read_last_successful_date()

if "isic_text" not in st.session_state:
    st.session_state["isic_text"] = load_default_isic_text()

with st.sidebar:
    st.header("Scrape Settings")
    all_pages = st.checkbox("All pages", value=False)
    page = st.number_input("Landing page (when All pages OFF)", min_value=1, value=1, step=1, disabled=all_pages)
    max_pages = st.number_input("Max pages to scan (safety)", min_value=1, value=10, step=1)
    max_tenders = st.number_input("Max tenders (before filtering)", min_value=1, value=30, step=1)
    concurrency = st.slider("Concurrency (details fetch)", min_value=1, max_value=10, value=6)

    st.divider()
    st.header("Published After")
    after_date = st.date_input("Include tenders published after:", value=st.session_state["after_date"])
    st.session_state["after_date"] = after_date

    cA, cB, cC = st.columns(3)
    with cA:
        if st.button("Last week"):
            st.session_state["after_date"] = date.today() - timedelta(days=7)
            st.rerun()
    with cB:
        if st.button("Last month"):
            st.session_state["after_date"] = date.today() - timedelta(days=30)
            st.rerun()
    with cC:
        if st.button("Since last run"):
            st.session_state["after_date"] = read_last_successful_date()
            st.rerun()

    include_missing_dates = st.checkbox("Include tenders with missing publish date", value=True)

    st.divider()
    st.header("Relevance Gates")
    margin_gate = st.slider("Margin gate (top1 - top2)", min_value=0.005, max_value=0.050, value=0.018, step=0.001)
    z_gate = st.slider("Z gate", min_value=0.1, max_value=3.0, value=0.75, step=0.05)
    st.divider()
    st.header("Interests")
    uploaded = st.file_uploader("Upload interests.json", type=["json"], key="upload_interests_json")
    if uploaded:
        st.session_state["interests_text"] = uploaded.read().decode("utf-8", errors="ignore")

    try:
        _io = json.loads(st.session_state.get("interests_text", "{}") or "{}")
        _interest_count = len(_io) if isinstance(_io, dict) else 0
    except Exception:
        _interest_count = 0
    st.caption(f"Loaded interests: **{_interest_count}**")

    cI1, cI2 = st.columns([1, 1])
    with cI1:
        if HAS_POPOVER:
            with st.popover("🧾 Edit Interests (table)", use_container_width=True):
                render_interests_editor_ui()
        else:
            with st.expander("🧾 Edit Interests (table)", expanded=False):
                render_interests_editor_ui()
    with cI2:
        st.download_button(
            "⬇️ Download interests.json",
            data=(st.session_state.get("interests_text", "") or "").encode("utf-8"),
            file_name="interests.json",
            mime="application/json",
            use_container_width=True,
        )

    st.divider()
    st.header("ISIC Matching (Activities list)")
    st.caption("Relevance = Embedding OR ISIC match. ISIC code is treated as TEXT (leading zeros preserved).")

    isic_up = st.file_uploader("Upload isic-of-interests.json", type=["json"], key="upload_isic_json")
    if isic_up:
        st.session_state["isic_text"] = isic_up.read().decode("utf-8", errors="ignore")

    try:
        _cfg = parse_isic_config(st.session_state.get("isic_text", "[]") or "[]")
        _isic_count = len(_cfg)
    except Exception:
        _isic_count = 0
    st.caption(f"Loaded ISIC rows: **{_isic_count}**")

    cS1, cS2 = st.columns([1, 1])
    with cS1:
        if HAS_POPOVER:
            with st.popover("🧾 Edit ISIC list (table)", use_container_width=True):
                render_isic_editor_ui()
        else:
            with st.expander("🧾 Edit ISIC list (table)", expanded=False):
                render_isic_editor_ui()
    with cS2:
        st.download_button(
            "⬇️ Download isic-of-interests.json",
            data=(st.session_state.get("isic_text", "[]") or "[]").encode("utf-8"),
            file_name="isic-of-interests.json",
            mime="application/json",
            use_container_width=True,
        )

colA, colB = st.columns([1, 1])
with colA:
    run_btn = st.button("🚀 Scrape & Score", type="primary")
with colB:
    st.write("Embedding model: `intfloat/multilingual-e5-large` (FastEmbed, cached locally)")

if run_btn:
    # Validate interests JSON
    try:
        interests_obj = json.loads(st.session_state["interests_text"])
        if not isinstance(interests_obj, dict) or not interests_obj:
            raise ValueError("Interests must be a non-empty JSON object.")
    except Exception as e:
        st.error(f"Invalid interests JSON: {e}")
        st.stop()

    with st.spinner("Initializing embeddings (first run may download model files)..."):
        interest_vecs = build_interest_vectors_cached(st.session_state["interests_text"])

    with st.spinner("Scraping Monqasat and loading tender details..."):
        # Decide how many pages to scan
        # - If user checks All Pages: start from page 1 for N pages
        # - Otherwise: start from selected page for N pages (NOT just 1)
        start_page = 1 if all_pages else int(page)
        pages_to_scan = int(max_pages)  # always respect the UI number

        raw_items = run_async(
            run_pipeline_pages(
                start_page=start_page,
                max_pages=pages_to_scan,
                max_tenders_total=int(max_tenders),
                concurrency=int(concurrency),
            )
        )

    # Filter by publish date BEFORE scoring (faster & more relevant)
    filtered = []
    for it in raw_items:
        pub_s = (it.get("landing") or {}).get("publish_date")
        pub_d = parse_date_any(pub_s)
        if pub_d is None:
            if include_missing_dates:
                filtered.append(it)
        else:
            if pub_d > st.session_state["after_date"]:
                filtered.append(it)

    # Parse ISIC config once per run
    try:
        isic_cfg = parse_isic_config(st.session_state.get("isic_text", "[]"))
    except Exception:
        isic_cfg = []

    items_out = []
    with st.spinner("Scoring relevance..."):
        prog = st.progress(0)
        total = max(1, len(filtered))
        for i, it in enumerate(filtered, start=1):
            landing = it["landing"]
            details = it["details"]

            subject = (details.get("subject") if details else None) or landing.get("title", "")
            brief = (details.get("brief_description") if details else None) or ""
            summary = summarize_topic(subject=subject, brief_desc=brief)

            tender_text = " ".join([
                landing.get("title") or "",
                landing.get("ministry") or "",
                landing.get("requested_sector_type") or "",
                (details.get("subject") or "") if details else "",
                (details.get("brief_description") or "") if details else "",
                ((details.get("general_conditions_text") or "")[:1500]) if details else "",
            ])

            score, breakdown, debug, explanation = semantic_relevance_standout(
                interest_vecs,
                tender_text,
                margin_gate=float(margin_gate),
                z_gate=float(z_gate),
            )

            # ISIC matching (from tender details -> activities list)
            tender_acts = []
            try:
                tender_acts = (details.get("activities") if details else None) or []
            except Exception:
                tender_acts = []

            isic_res = match_isic_activities(tender_acts, isic_cfg) if isic_cfg else {"match_count": 0, "matched_codes": [], "matched_items": []}
            isic_match_count = int(isic_res.get("match_count") or 0)
            isic_matched_codes = isic_res.get("matched_codes") or []
            isic_matched_items = isic_res.get("matched_items") or []

            # --- FIX: If ISIC match exists, do NOT keep label/score as Irrelevant just because embedding gated out ---
            if isic_match_count > 0:
                # Always expose ISIC info in debug
                debug = dict(debug or {})
                debug["isic_match_count"] = isic_match_count
                debug["isic_matched_codes"] = list(isic_matched_codes)

                # If embedding score is 0 (gated), override the displayed score/label to be relevant
                if int(score or 0) <= 0:
                    # Treat ISIC match as a strong relevance signal
                    # 1 match -> 70, 2 -> 80, 3 -> 90, 4+ -> 100
                    score = min(100, 70 + 10 * (max(0, isic_match_count - 1)))
                    if score >= 70:
                        debug["label"] = "Relevant (High)"
                    elif score >= 40:
                        debug["label"] = "Relevant (Medium)"
                    else:
                        debug["label"] = "Relevant (Low)"

                    explanation = f"{explanation} | ISIC match ({isic_match_count}) => override to {debug['label']} (score={score})"

            overall_relevant = (int(score or 0) > 0) or (isic_match_count > 0)

            embedding_match = (int(score or 0) > 0)
            isic_match = (isic_match_count > 0)
            if embedding_match and isic_match:
                overall_reason = "Embedding + ISIC"
            elif embedding_match:
                overall_reason = "Embedding"
            elif isic_match:
                overall_reason = "ISIC"
            else:
                overall_reason = "None"

            items_out.append({
                "landing": landing,
                "details": details,
                "topic_summary": summary,
                "relevance_score": score,
                "isic_match_count": isic_match_count,
                "isic_matched_codes": isic_matched_codes,
                "isic_matched_items": isic_matched_items,
                "overall_relevant": overall_relevant,
                "overall_relevance_reason": overall_reason,
                "relevance_breakdown": breakdown,
                "relevance_debug": debug,
                "relevance_explanation": explanation,
                "scraped_at_utc": now_utc_iso(),
            })

            prog.progress(i / total)

    st.session_state["items_out"] = items_out

    # Update "last successful run" date only when we produced output
    if items_out is not None:
        write_last_successful_date(date.today())


# Render results if available
items_out = st.session_state.get("items_out")
if items_out:
    st.success(f"Done. Scored {len(items_out)} tenders (after date filtering).")

    relevant = [x for x in items_out if (x.get("overall_relevant") is True) or ((x.get("relevance_score", 0) or 0) > 0)]
    irrelevant = [x for x in items_out if not ((x.get("overall_relevant") is True) or ((x.get("relevance_score", 0) or 0) > 0))]

    json_bytes = json.dumps(items_out, ensure_ascii=False, indent=2).encode("utf-8")
    csv_bytes = to_csv_bytes(items_out)
    xlsx_bytes = to_excel_bytes(items_out)

    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button("⬇️ JSON", data=json_bytes, file_name="tenders.json", mime="application/json")
    with d2:
        st.download_button("⬇️ CSV", data=csv_bytes, file_name="tenders.csv", mime="text/csv")
    with d3:
        st.download_button("⬇️ Excel", data=xlsx_bytes, file_name="tenders.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.subheader(f"✅ Relevant tenders ({len(relevant)})")
    if relevant:
        components.html(
            build_html_table(relevant, f"✅ Relevant tenders ({len(relevant)})"),
            height=520,
            scrolling=True
        )

        with st.expander("🔎 Relevant details (expanders)"):
            for it in sorted(relevant, key=lambda t: t.get("relevance_score", 0), reverse=True):
                landing = it["landing"]
                dbg = it.get("relevance_debug") or {}
                title = landing.get("title", "")
                tender_no = landing.get("tender_no", "")
                score = it.get("relevance_score", 0)
                best = dbg.get("best_interest", "")
                reason = it.get("overall_relevance_reason", "")
                with st.expander(f"[{score}] {tender_no} | {reason} | {best} | {title}"):
                    st.write(it.get("relevance_explanation", ""))
                    st.write(f"Topic summary: {it.get('topic_summary','')}")
                    st.write(f"Details URL: {landing.get('details_url','')}")
                    st.write(f"ISIC match count: {it.get('isic_match_count',0)}")
                    st.write(f"ISIC matched codes: {', '.join(it.get('isic_matched_codes') or [])}")
                    st.json(it.get("relevance_breakdown", []))
    else:
        st.info("No relevant tenders found under current filters/gates.")

    st.divider()

    st.subheader(f"🚫 Irrelevant / unclear tenders ({len(irrelevant)})")
    if irrelevant:
        components.html(
            build_html_table(irrelevant, f"🚫 Irrelevant / unclear tenders ({len(irrelevant)})"),
            height=520,
            scrolling=True
        )
    else:
        st.info("No irrelevant tenders under current filters.")